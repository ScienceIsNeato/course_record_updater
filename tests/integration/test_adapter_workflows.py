"""
Integration tests for adapter-based import/export workflows

Tests complete end-to-end workflows for different user roles using the
adapter registry system for bidirectional data flow.
"""

import json
import tempfile
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook

from adapters.adapter_registry import get_adapter_registry
from database_factory import get_database_service
from database_service import (
    create_default_cei_institution,
    get_active_terms,
    get_all_courses,
    get_all_users,
)
from export_service import ExportConfig, ExportService
from import_service import ConflictStrategy, ImportService


@pytest.mark.integration
class TestAdapterWorkflows:
    """Integration tests for complete adapter workflows."""

    def setup_method(self):
        """Set up test environment."""
        self.registry = get_adapter_registry()

        # Ensure CEI institution exists
        self.institution_id = create_default_cei_institution()

        # Initialize services with institution ID
        self.import_service = ImportService(self.institution_id)
        self.export_service = ExportService()

    def create_test_excel_file(self, file_path: str, data: list) -> None:
        """Create a test Excel file with CEI format data."""
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Test Data"

        # Headers for CEI format
        headers = [
            "course",
            "section",
            "effterm_c",
            "students",
            "Faculty Name",
            "email",
        ]

        # Write headers
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)

        # Write data rows
        for row, record in enumerate(data, 2):
            worksheet.cell(row=row, column=1, value=record.get("course", ""))
            worksheet.cell(row=row, column=2, value=record.get("section", ""))
            worksheet.cell(row=row, column=3, value=record.get("effterm_c", ""))
            worksheet.cell(row=row, column=4, value=record.get("students", ""))
            worksheet.cell(row=row, column=5, value=record.get("Faculty Name", ""))
            worksheet.cell(row=row, column=6, value=record.get("email", ""))

        workbook.save(file_path)

    def test_site_admin_full_import_export_workflow(self):
        """Test complete import/export workflow for site admin."""
        # Create test data (using correct term format FA2024)
        test_data = [
            {
                "course": "MATH-101",
                "section": "01",
                "effterm_c": "FA2024",
                "students": "25",
                "Faculty Name": "Dr. John Smith",
                "email": "john.smith@cei.edu",
            },
            {
                "course": "ENG-201",
                "section": "01",
                "effterm_c": "FA2024",
                "students": "30",
                "Faculty Name": "Prof. Jane Doe",
                "email": "jane.doe@cei.edu",
            },
        ]

        with tempfile.TemporaryDirectory() as tmp_dir:
            # Step 1: Create test Excel file
            input_file = Path(tmp_dir) / "test_import.xlsx"
            self.create_test_excel_file(str(input_file), test_data)

            # Step 2: Import data using CEI adapter
            import_result = self.import_service.import_excel_file(
                file_path=str(input_file),
                conflict_strategy=ConflictStrategy.USE_THEIRS,
                dry_run=False,
                adapter_id="cei_excel_format_v1",
            )

            # Verify import success
            assert import_result.success is True
            assert import_result.records_processed > 0

            # Step 3: Verify data was imported
            courses = get_all_courses(self.institution_id)
            users = get_all_users(self.institution_id)
            terms = get_active_terms(self.institution_id)

            assert len(courses) >= 2  # At least our test courses
            assert len(users) >= 2  # At least our test users
            assert len(terms) >= 1  # At least one term

            # Step 4: Export data using same adapter
            export_config = ExportConfig(
                institution_id=self.institution_id, adapter_id="cei_excel_format_v1"
            )

            export_file = Path(tmp_dir) / "test_export.xlsx"
            export_result = self.export_service.export_data(export_config, export_file)

            # Verify export success
            assert export_result.success is True
            assert export_result.records_exported > 0
            assert export_file.exists()

            # Step 5: Verify exported file structure
            exported_df = pd.read_excel(export_file)
            assert "course" in exported_df.columns
            assert "Faculty Name" in exported_df.columns
            assert "email" in exported_df.columns
            assert len(exported_df) > 0

    def test_institution_admin_adapter_access(self):
        """Test that institution admin can only access their institution's adapters."""
        # Use the actual CEI institution ID
        user = {
            "role": "institution_admin",
            "institution_id": self.institution_id,  # Use the actual CEI institution ID
        }

        # Should have access to CEI adapter
        has_access, message = self.export_service.validate_export_access(
            user, "cei_excel_format_v1"
        )
        assert has_access is True
        assert message == "Access granted"

        # Should not have access to non-existent adapter
        has_access, message = self.export_service.validate_export_access(
            user, "other_institution_adapter"
        )
        assert has_access is False
        assert "Access denied" in message

    def test_instructor_export_restrictions(self):
        """Test that instructors cannot export data."""
        user = {"role": "instructor", "institution_id": self.institution_id}

        # Should not have access to any export functionality
        has_access, message = self.export_service.validate_export_access(
            user, "cei_excel_format_v1"
        )
        assert has_access is False
        assert "Access denied" in message

    def test_adapter_registry_discovery(self):
        """Test that adapter registry correctly discovers available adapters."""
        # Get all adapters
        adapters = self.registry.get_all_adapters()
        assert len(adapters) >= 1

        # Should find CEI adapter
        cei_adapter_found = any(
            adapter["id"] == "cei_excel_format_v1" for adapter in adapters
        )
        assert cei_adapter_found is True

        # Test getting specific adapter
        cei_adapter = self.registry.get_adapter_by_id("cei_excel_format_v1")
        assert cei_adapter is not None
        assert cei_adapter.supports_export() is True

    def test_roundtrip_data_consistency(self):
        """Test that import->export->import maintains data consistency."""
        # Original test data (using correct term format SP2024)
        original_data = [
            {
                "course": "PHYS-301",
                "section": "01",
                "effterm_c": "SP2024",
                "students": "15",
                "Faculty Name": "Dr. Albert Einstein",
                "email": "einstein@cei.edu",
            }
        ]

        with tempfile.TemporaryDirectory() as tmp_dir:
            # Step 1: Create original file
            original_file = Path(tmp_dir) / "original.xlsx"
            self.create_test_excel_file(str(original_file), original_data)

            # Step 2: Import original data
            import_result1 = self.import_service.import_excel_file(
                file_path=str(original_file),
                conflict_strategy=ConflictStrategy.USE_THEIRS,
                dry_run=False,
                adapter_id="cei_excel_format_v1",
            )
            assert import_result1.success is True

            # Step 3: Export the imported data
            export_config = ExportConfig(
                institution_id=self.institution_id, adapter_id="cei_excel_format_v1"
            )

            exported_file = Path(tmp_dir) / "exported.xlsx"
            export_result = self.export_service.export_data(
                export_config, exported_file
            )
            assert export_result.success is True

            # Step 4: Import the exported data (roundtrip)
            import_result2 = self.import_service.import_excel_file(
                file_path=str(exported_file),
                conflict_strategy=ConflictStrategy.USE_THEIRS,
                dry_run=False,
                adapter_id="cei_excel_format_v1",
            )
            assert import_result2.success is True

            # Step 5: Verify data consistency
            final_courses = get_all_courses(self.institution_id)
            final_users = get_all_users(self.institution_id)

            # Should have the same course
            physics_course = next(
                (c for c in final_courses if c.get("course_number") == "PHYS-301"), None
            )
            assert physics_course is not None

            # Should have the instructor
            einstein = next(
                (u for u in final_users if "Einstein" in u.get("last_name", "")), None
            )
            assert einstein is not None
            assert einstein.get("email") == "einstein@cei.edu"

    def test_adapter_error_handling(self):
        """Test adapter error handling for invalid files."""
        with tempfile.TemporaryDirectory() as tmp_dir:
            # Create invalid file (wrong format)
            invalid_file = Path(tmp_dir) / "invalid.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.cell(row=1, column=1, value="wrong_header")
            workbook.save(invalid_file)

            # Should fail compatibility check
            adapter = self.registry.get_adapter_by_id("cei_excel_format_v1")
            is_compatible, message = adapter.validate_file_compatibility(
                str(invalid_file)
            )
            assert is_compatible is False
            # The actual error message from the adapter
            assert (
                "Excel file is empty" in message
                or "Required columns not found" in message
            )

            # Import should fail gracefully
            import_result = self.import_service.import_excel_file(
                file_path=str(invalid_file),
                conflict_strategy=ConflictStrategy.USE_THEIRS,
                dry_run=False,
                adapter_id="cei_excel_format_v1",
            )
            assert import_result.success is False
            assert len(import_result.errors) > 0

    def test_multiple_adapter_support(self):
        """Test system support for multiple adapters (extensibility)."""
        # Get supported formats (returns dict mapping adapter_id -> formats)
        supported_formats = self.registry.get_supported_formats()
        assert "cei_excel_format_v1" in supported_formats
        assert ".xlsx" in supported_formats["cei_excel_format_v1"]

        # Test finding adapters by format
        xlsx_adapters = self.registry.find_adapters_for_format(".xlsx")
        assert len(xlsx_adapters) >= 1
        assert any(adapter["id"] == "cei_excel_format_v1" for adapter in xlsx_adapters)

        # Test role-based adapter filtering
        site_admin_adapters = self.registry.get_adapters_for_user(
            "site_admin", self.institution_id
        )
        # Use the actual CEI institution ID
        institution_admin_adapters = self.registry.get_adapters_for_user(
            "institution_admin", self.institution_id
        )
        instructor_adapters = self.registry.get_adapters_for_user(
            "instructor", self.institution_id
        )

        # Site admin should see all adapters
        assert len(site_admin_adapters) >= 1

        # Institution admin should see their institution's adapters
        assert len(institution_admin_adapters) >= 1

        # Instructors should see no adapters (no import/export permissions)
        assert len(instructor_adapters) == 0

    def test_generic_csv_adapter_export_and_parse_with_database(self):
        """
        Integration test: Generic CSV adapter export and parse with real database.

        Tests workflow:
        1. Create realistic data in database
        2. Export to ZIP using generic CSV adapter
        3. Verify ZIP structure and contents
        4. Parse ZIP back to verify round-trip data integrity
        """
        with tempfile.TemporaryDirectory() as tmp_dir:
            db = get_database_service()
            output_file = Path(tmp_dir) / "generic_csv_export.zip"

            # Step 1: Create test data
            institution_id = db.create_institution(
                {
                    "name": "CSV Test University",
                    "short_name": "CSVTU",
                    "admin_email": "admin@csvtu.edu",
                }
            )

            program_id = db.create_program(
                {
                    "name": "Computer Science",
                    "short_name": "CS",
                    "institution_id": institution_id,
                    "is_active": True,
                }
            )

            user1_id = db.create_user(
                {
                    "email": "prof1@csvtu.edu",
                    "first_name": "John",
                    "last_name": "Doe",
                    "role": "instructor",
                    "institution_id": institution_id,
                }
            )

            user2_id = db.create_user(
                {
                    "email": "prof2@csvtu.edu",
                    "first_name": "Jane",
                    "last_name": "Smith",
                    "role": "instructor",
                    "institution_id": institution_id,
                }
            )

            course_id = db.create_course(
                {
                    "course_number": "CS101",
                    "course_title": "Intro to CS",
                    "department": "CS",
                    "credit_hours": 3,
                    "institution_id": institution_id,
                }
            )

            db.add_course_to_program(course_id, program_id)

            term_id = db.create_term(
                {
                    "term_name": "FA2024",
                    "name": "Fall 2024",
                    "active": True,
                    "institution_id": institution_id,
                }
            )

            # Step 2: Export using generic CSV adapter
            adapter = self.registry.get_adapter_by_id("generic_csv_v1")
            assert adapter is not None

            config = ExportConfig(
                institution_id=institution_id,
                adapter_id="generic_csv_v1",
                export_view="standard",
                include_metadata=True,
                output_format="zip",
            )

            export_result = self.export_service.export_data(config, str(output_file))
            assert export_result.success is True
            assert export_result.records_exported > 0
            assert output_file.exists()

            # Step 3: Verify ZIP structure
            import zipfile

            with zipfile.ZipFile(output_file, "r") as zf:
                filenames = zf.namelist()
                assert "manifest.json" in filenames
                assert "institutions.csv" in filenames
                assert "users.csv" in filenames
                assert "courses.csv" in filenames
                assert "terms.csv" in filenames

                # Verify manifest
                manifest = json.loads(zf.read("manifest.json"))
                assert manifest["format_version"] == "1.0"
                assert "entity_counts" in manifest
                assert manifest["entity_counts"]["users"] == 2
                assert manifest["entity_counts"]["courses"] == 1

            # Step 4: Parse ZIP to verify round-trip integrity
            parsed_data = adapter.parse_file(str(output_file), {})

            # Verify parsed data structure
            assert "institutions" in parsed_data
            assert "users" in parsed_data
            assert "courses" in parsed_data
            assert "terms" in parsed_data

            # Verify counts match
            assert len(parsed_data["users"]) == 2
            assert len(parsed_data["courses"]) == 1
            assert len(parsed_data["terms"]) == 1

            # Verify data integrity
            prof1 = next(
                (u for u in parsed_data["users"] if u["email"] == "prof1@csvtu.edu"),
                None,
            )
            assert prof1 is not None
            assert prof1["first_name"] == "John"

            cs101 = parsed_data["courses"][0]
            assert cs101["course_number"] == "CS101"
            assert cs101["credit_hours"] == "3"  # String from CSV
