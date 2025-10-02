"""
CEI Excel Import Adapter

This adapter handles the specific Excel format used by CEI (College of Eastern Idaho).
It contains all CEI-specific parsing logic, column mappings, and data transformations.

This keeps CEI-specific logic separate from the generic import system.
"""

from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd

from models import validate_course_number

from .file_base_adapter import FileBaseAdapter, FileCompatibilityError

# Constants for repeated strings
FACULTY_NAME_COLUMN = "Faculty Name"
ENROLLED_STUDENTS_COLUMN = "Enrolled Students"
XLSX_EXTENSION = ".xlsx"
XLS_EXTENSION = ".xls"


def validate_cei_term_name(term_name: str) -> bool:
    """
    Validate CEI-specific term name formats

    Supports:
    - Standard format: "2024 Fall", "2024 Spring", etc.
    - CEI abbreviated format: "2024FA", "2024SP", "2024SU", "2024WI"
    """
    # Handle space-separated format: "2024 Fall"
    parts = term_name.split()
    if len(parts) == 2 and parts[0].isdigit() and len(parts[0]) == 4:
        season = parts[1].lower()
        return season in ["fall", "spring", "summer", "winter"]

    # Handle CEI abbreviated format: "FA2024", "SP2025", "SU2023", "WI2026"
    if len(term_name) == 6 and term_name[2:].isdigit():
        season = term_name[:2].upper()
        return season in ["FA", "SP", "SU", "WI"]  # Fall, Spring, Summer, Winter

    return False


def parse_cei_term(effterm_c: str) -> Tuple[str, str]:
    """
    Parse CEI's effterm_c format into year and season.

    Args:
        effterm_c: Term code like 'FA2024', 'SP2025', etc.

    Returns:
        Tuple of (year, season) where season is the full name

    Example:
        parse_cei_term('FA2024') -> ('2024', 'Fall')
    """
    if not effterm_c or len(effterm_c) != 6:
        raise ValueError(f"Invalid effterm_c format: {effterm_c}")

    season_code = effterm_c[:2].upper()
    year = effterm_c[2:]

    season_map = {"FA": "Fall", "SP": "Spring", "SU": "Summer", "WI": "Winter"}

    if season_code not in season_map:
        raise ValueError(f"Invalid season code: {season_code}")

    return year, season_map[season_code]


def _extract_course_data(
    row: pd.Series, institution_id: str
) -> Optional[Dict[str, Any]]:
    """Extract course information from row."""
    if "course" not in row or pd.isna(row["course"]):
        return None

    course_number = str(row.get("course", ""))
    if not validate_course_number(course_number):
        return None

    return {
        "course_number": course_number,
        "course_title": f"Course {course_number}",  # CEI file doesn't have course titles
        "department": _extract_department_from_course(course_number),
        "credit_hours": 3,  # Default, CEI file doesn't have credit hours
        "institution_id": institution_id,
    }


def _extract_user_data(
    row: pd.Series, institution_id: str, course_data: Optional[Dict[str, Any]]
) -> Optional[Dict[str, Any]]:
    """Extract instructor information from row."""

    # Handle two different input formats:
    # 1. Files with email column (test data format)
    # 2. Files with Faculty Name column (original format)

    if "email" in row and not pd.isna(row["email"]):
        # Format 1: Has email column - extract name from email if needed
        email = str(row["email"]).strip()
        if not email:
            return None

        # Try to extract name from email or use placeholder
        first_name, last_name = _extract_name_from_email(email)

        return {
            "email": email,
            "first_name": first_name,
            "last_name": last_name,
            "role": "instructor",
            "department": course_data.get("department") if course_data else None,
            "institution_id": institution_id,
            "account_status": "imported",
            "active_user": False,
        }

    elif FACULTY_NAME_COLUMN in row and not pd.isna(row[FACULTY_NAME_COLUMN]):
        # Format 2: Has Faculty Name but NO email - this is a problem!
        # We cannot and should not generate fake emails
        instructor_name = str(row[FACULTY_NAME_COLUMN])
        first_name, last_name = _parse_name(instructor_name)

        return {
            "email": None,  # No email available - must be added manually later
            "first_name": first_name,
            "last_name": last_name,
            "role": "instructor",
            "department": course_data.get("department") if course_data else None,
            "institution_id": institution_id,
            "account_status": "needs_email",  # Flag that email is required
            "active_user": False,
        }

    return None  # No instructor information available


def _extract_term_data(row: pd.Series, institution_id: str) -> Optional[Dict[str, Any]]:
    """Extract term information from row."""

    # Handle two different input formats:
    # 1. Files with effterm_c column (original format like "FA2024")
    # 2. Files with Term column (standard format like "2024 Fall")

    if "effterm_c" in row and not pd.isna(row["effterm_c"]):
        # Format 1: CEI abbreviated format
        effterm_c = str(row["effterm_c"]).strip()
        if not effterm_c:
            return None

        try:
            year, season = parse_cei_term(effterm_c)
            return {
                "name": f"{season} {year}",
                "term_name": f"{season} {year}",  # Add term_name for import service
                "year": int(year),
                "season": season,
                "institution_id": institution_id,
                "is_active": True,
                "start_date": f"{year}-01-01",  # Default start date
                "end_date": f"{year}-12-31",  # Default end date
            }
        except ValueError as e:
            print(f"Warning: Could not parse term '{effterm_c}': {e}")
            return None

    elif "Term" in row and not pd.isna(row["Term"]):
        # Format 2: Standard format like "2024 Fall"
        term_name = str(row["Term"]).strip()
        if not term_name:
            return None

        try:
            # Parse standard format "YYYY Season"
            parts = term_name.split()
            if len(parts) >= 2 and parts[0].isdigit():
                parsed_year = int(parts[0])
                parsed_season = parts[1].title()  # Fall, Spring, etc.

                return {
                    "name": term_name,
                    "term_name": term_name,  # Add term_name for import service
                    "year": parsed_year,
                    "season": parsed_season,
                    "institution_id": institution_id,
                    "is_active": True,
                    "start_date": f"{parsed_year}-01-01",  # Default start date
                    "end_date": f"{parsed_year}-12-31",  # Default end date
                }
        except (ValueError, IndexError) as e:
            print(f"Warning: Could not parse term '{term_name}': {e}")
            return None

    return None


def _extract_offering_data(
    course_data: Optional[Dict[str, Any]],
    term_data: Optional[Dict[str, Any]],
    user_data: Optional[Dict[str, Any]],
    institution_id: str,
) -> Optional[Dict[str, Any]]:
    """Extract offering information from course and term data."""
    if not course_data or not term_data:
        return None

    return {
        "course_number": course_data["course_number"],
        "term_name": term_data["name"],
        "institution_id": institution_id,
        "instructor_email": user_data.get("email") if user_data else None,
        "is_active": True,
    }


def _extract_section_data(
    row: pd.Series,
    course_data: Optional[Dict[str, Any]],
    term_data: Optional[Dict[str, Any]],
    user_data: Optional[Dict[str, Any]],
    institution_id: str,
) -> Optional[Dict[str, Any]]:
    """Extract section information from row and related data."""
    if not course_data or not term_data:
        return None

    # Check for either "Enrolled Students" (original format) or "students" (test format)
    student_count = None
    if ENROLLED_STUDENTS_COLUMN in row and not pd.isna(row[ENROLLED_STUDENTS_COLUMN]):
        try:
            student_count = int(row[ENROLLED_STUDENTS_COLUMN])
        except (ValueError, TypeError):
            pass
    elif "students" in row and not pd.isna(row["students"]):
        try:
            student_count = int(row["students"])
        except (ValueError, TypeError):
            pass

    if student_count is None:
        return None

    return {
        "course_number": course_data["course_number"],
        "term_name": term_data["name"],
        "section_number": "001",  # CEI doesn't provide section numbers
        "instructor_email": user_data.get("email") if user_data else None,
        "student_count": student_count,
        "institution_id": institution_id,
        "status": "active",
    }


def parse_cei_excel_row(
    row: pd.Series, institution_id: str
) -> Dict[str, Optional[Dict[str, Any]]]:
    """
    Parse a single row from CEI's Excel format.

    Args:
        row: Pandas Series representing one row
        institution_id: The institution ID to assign to entities

    Returns:
        Dictionary with entity types and their data
    """
    try:
        # Extract all entity data using focused helper functions
        course_data = _extract_course_data(row, institution_id)
        user_data = _extract_user_data(row, institution_id, course_data)
        term_data = _extract_term_data(row, institution_id)
        offering_data = _extract_offering_data(
            course_data, term_data, user_data, institution_id
        )
        section_data = _extract_section_data(
            row, course_data, term_data, user_data, institution_id
        )

        return {
            "course": course_data,
            "user": user_data,
            "term": term_data,
            "offering": offering_data,
            "section": section_data,
        }

    except Exception as e:
        # Log error and return empty entities to continue processing
        print(f"Error parsing CEI Excel row: {str(e)}")
        return {
            "course": None,
            "user": None,
            "term": None,
            "offering": None,
            "section": None,
        }


def _extract_department_from_course(course_number: str) -> str:
    """Extract department from course number (e.g., 'MATH-101' -> 'Mathematics')."""
    # CEI-specific department mapping
    department_map = {
        "MATH": "Mathematics",
        "ENG": "English",
        "HIST": "History",
        "SCI": "Science",
        "BUS": "Business",
        "ART": "Art",
        "MUS": "Music",
        "PE": "Physical Education",
        "COMP": "Computer Science",
        "BIO": "Biology",
        "CHEM": "Chemistry",
        "PHYS": "Physics",
        "SOC": "Sociology",
        "PSYC": "Psychology",
        "ECON": "Economics",
        "POLS": "Political Science",
        "PHIL": "Philosophy",
        "REL": "Religious Studies",
        "FOR": "Foreign Language",
        "NURS": "Nursing",
        "EDU": "Education",
    }

    if "-" in course_number:
        prefix = course_number.split("-")[0].upper()
        return department_map.get(prefix, prefix)
    return "General Studies"


def _parse_name(full_name: str) -> Tuple[str, str]:
    """Parse full name into first and last name."""
    if not full_name or not full_name.strip():
        return "Unknown", "Instructor"

    parts = full_name.strip().split()
    if len(parts) == 1:
        return parts[0], "Name"
    elif len(parts) == 2:
        return parts[0], parts[1]
    else:
        # More than 2 parts, assume first is first name, rest is last name
        return parts[0], " ".join(parts[1:])


def _extract_name_from_email(email: str) -> Tuple[str, str]:
    """Extract first and last name from email address."""
    if not email or "@" not in email:
        return "Unknown", "Instructor"

    local_part = email.split("@")[0]

    # Handle common email formats
    if "." in local_part:
        # firstname.lastname format
        parts = local_part.split(".")
        first_name = parts[0].title()
        last_name = ".".join(parts[1:]).title()
    else:
        # Single name or other format - use as last name
        first_name = "Unknown"
        last_name = local_part.title()

    return first_name, last_name


class CEIExcelAdapter(FileBaseAdapter):
    """
    Adapter for CEI's Excel format with automatic compatibility detection.

    Handles CEI's specific Excel format with dual format support:
    1. Original format with Faculty Name and effterm_c columns
    2. Test format with email and Term columns

    Automatically detects data types and validates file compatibility.
    """

    SUPPORTED_EXTENSIONS = [XLSX_EXTENSION, XLS_EXTENSION]
    MAX_FILE_SIZE_MB = 500
    MAX_RECORDS_TO_PROCESS = 500000

    def __init__(self):
        """Initialize CEI Excel adapter with format specifications."""
        super().__init__()

        # Adapter identification
        self.adapter_id = "cei_excel_format_v1"
        self.institution_id = "cei_institution_id"

        # Required columns for original format (flexible student column)
        self.original_format_required = ["course", FACULTY_NAME_COLUMN, "effterm_c"]
        self.original_format_student_columns = [
            ENROLLED_STUDENTS_COLUMN,
            "students",
        ]  # Either works

        # Required columns for test format
        self.test_format_columns = ["course", "email", "Term", "students"]

        # Required columns for hybrid format (real CEI data with both email and Term)
        self.hybrid_format_required = ["course", "email", "Term"]
        self.hybrid_format_student_columns = [
            ENROLLED_STUDENTS_COLUMN,
            "students",
        ]  # Either works

        # Optional columns that may be present
        self.optional_columns = ["Course Title", "Department", "Credits"]

    def validate_file_compatibility(self, file_path: str) -> Tuple[bool, str]:
        """
        Validate that the file is compatible with CEI Excel format.

        Checks for:
        1. Excel file format (.xlsx or .xls)
        2. Required column structure (either original or test format)
        3. Valid data patterns in key columns

        Returns:
            Tuple[bool, str]: (is_compatible, message)
        """
        try:
            # Basic file validation
            basic_check = self._validate_basic_file_properties(file_path)
            if not basic_check[0]:
                return basic_check

            # Read and validate Excel structure
            df = self._read_excel_sample(file_path)
            if isinstance(df, tuple):  # Error case
                return df

            # Check format compatibility
            format_check = self._check_format_compatibility(df)
            if not format_check[0]:
                return format_check

            # Validate data patterns
            validation_errors = self._validate_data_patterns(df, format_check[1])
            if validation_errors:
                return False, f"Data validation failed: {'; '.join(validation_errors)}"

            # Success message
            record_count = len(df)
            return True, (
                f"File compatible with CEI Excel format ({format_check[1]}). "
                f"Found {record_count} sample records."
            )

        except Exception as e:
            return False, f"Error validating file: {str(e)}"

    def _validate_basic_file_properties(self, file_path: str) -> Tuple[bool, str]:
        """Validate basic file properties: extension, existence, size."""
        import os

        # Check file extension
        if not file_path.lower().endswith((XLSX_EXTENSION, XLS_EXTENSION)):
            return False, "Unsupported file extension. Expected .xlsx or .xls"

        # Check if file exists
        if not os.path.exists(file_path):
            return False, f"File not found: {file_path}"

        # Check file size
        file_size_mb = os.path.getsize(file_path) / (1024 * 1024)
        if file_size_mb > self.MAX_FILE_SIZE_MB:
            return (
                False,
                f"File too large ({file_size_mb:.1f}MB). Maximum allowed: {self.MAX_FILE_SIZE_MB}MB",
            )

        return True, "Basic file validation passed"

    def _read_excel_sample(self, file_path: str):
        """Read Excel file sample for validation."""
        try:
            df = pd.read_excel(file_path, nrows=10)  # Sample first 10 rows
            if df.empty:
                return False, "Excel file is empty"
            return df
        except Exception as e:
            return False, f"Cannot read Excel file: {str(e)}"

    def _check_format_compatibility(self, df) -> Tuple[bool, str]:
        """Check if DataFrame matches any supported CEI format."""
        # Check original format
        has_original_required = all(
            col in df.columns for col in self.original_format_required
        )
        has_original_student_col = any(
            col in df.columns for col in self.original_format_student_columns
        )
        has_original_format = has_original_required and has_original_student_col

        # Check test format
        has_test_format = all(col in df.columns for col in self.test_format_columns)

        # Check hybrid format
        has_hybrid_required = all(
            col in df.columns for col in self.hybrid_format_required
        )
        has_hybrid_student_col = any(
            col in df.columns for col in self.hybrid_format_student_columns
        )
        has_hybrid_format = has_hybrid_required and has_hybrid_student_col

        if has_original_format:
            return True, "original"
        elif has_test_format:
            return True, "test"
        elif has_hybrid_format:
            return True, "hybrid"
        else:
            return False, self._build_format_error_message(df)

    def _build_format_error_message(self, df) -> str:
        """Build detailed error message for format mismatch."""
        missing_original_required = [
            col for col in self.original_format_required if col not in df.columns
        ]
        has_original_student_col = any(
            col in df.columns for col in self.original_format_student_columns
        )
        missing_original_student = (
            "students column" if not has_original_student_col else None
        )
        missing_test = [
            col for col in self.test_format_columns if col not in df.columns
        ]

        missing_hybrid_required = [
            col for col in self.hybrid_format_required if col not in df.columns
        ]
        has_hybrid_student_col = any(
            col in df.columns for col in self.hybrid_format_student_columns
        )
        missing_hybrid_student = (
            "students column" if not has_hybrid_student_col else None
        )

        missing_original = missing_original_required + (
            [missing_original_student] if missing_original_student else []
        )
        missing_hybrid = missing_hybrid_required + (
            [missing_hybrid_student] if missing_hybrid_student else []
        )

        return (
            f"File doesn't match CEI format. "
            f"Missing for original format: {missing_original}. "
            f"Missing for test format: {missing_test}. "
            f"Missing for hybrid format: {missing_hybrid}."
        )

    def _validate_data_patterns(self, df, format_type: str) -> List[str]:
        """Validate data patterns in key columns."""
        validation_errors = []

        # Validate course column
        course_error = self._validate_course_column(df)
        if course_error:
            validation_errors.append(course_error)

        # Validate term column for original format
        if format_type == "original":
            term_error = self._validate_term_column(df)
            if term_error:
                validation_errors.append(term_error)

        # Validate student count column
        student_error = self._validate_student_count_column(df)
        if student_error:
            validation_errors.append(student_error)

        return validation_errors

    def _validate_course_column(self, df) -> Optional[str]:
        """Validate course column data patterns."""
        if "course" not in df.columns:
            return None

        sample_courses = df["course"].dropna().head(3)
        invalid_courses = [
            course
            for course in sample_courses
            if not validate_course_number(str(course))
        ]
        if len(invalid_courses) == len(sample_courses) and len(sample_courses) > 0:
            return "No valid course numbers found"
        return None

    def _validate_term_column(self, df) -> Optional[str]:
        """Validate term column data patterns for original format."""
        if "effterm_c" not in df.columns:
            return None

        sample_terms = df["effterm_c"].dropna().head(3)
        invalid_terms = [
            term for term in sample_terms if not validate_cei_term_name(str(term))
        ]
        if len(invalid_terms) == len(sample_terms) and len(sample_terms) > 0:
            return "No valid term codes found (expected format: FA2024, SP2025)"
        return None

    def _validate_student_count_column(self, df) -> Optional[str]:
        """Validate student count column data patterns."""
        student_col = None
        if ENROLLED_STUDENTS_COLUMN in df.columns:
            student_col = ENROLLED_STUDENTS_COLUMN
        elif "students" in df.columns:
            student_col = "students"

        if not student_col:
            return None

        sample_students = df[student_col].dropna().head(3)
        try:
            numeric_counts = [int(x) for x in sample_students if pd.notna(x)]
            if not numeric_counts and len(sample_students) > 0:
                return f"Student count column ({student_col}) contains no valid numbers"
        except (ValueError, TypeError):
            return f"Student count column ({student_col}) contains invalid data"
        return None

    def detect_data_types(self, file_path: str) -> List[str]:
        """
        Automatically detect what types of data are present in the file.

        CEI Excel files typically contain:
        - courses: Course information
        - faculty: Instructor information
        - terms: Academic term data
        - sections: Course section details
        """
        try:
            df = pd.read_excel(file_path, nrows=50)  # Sample for analysis
            detected_types = []

            # Check for course data
            if "course" in df.columns and not df["course"].isna().all():
                detected_types.append("courses")

            # Check for faculty data
            if (
                FACULTY_NAME_COLUMN in df.columns
                and not df[FACULTY_NAME_COLUMN].isna().all()
            ) or ("email" in df.columns and not df["email"].isna().all()):
                detected_types.append("faculty")

            # Check for term data
            if ("effterm_c" in df.columns and not df["effterm_c"].isna().all()) or (
                "Term" in df.columns and not df["Term"].isna().all()
            ):
                detected_types.append("terms")

            # Check for section data (student counts indicate sections)
            if (
                ENROLLED_STUDENTS_COLUMN in df.columns
                and not df[ENROLLED_STUDENTS_COLUMN].isna().all()
            ) or ("students" in df.columns and not df["students"].isna().all()):
                detected_types.append("sections")

            return detected_types

        except Exception:
            # If we can't read the file, return empty list
            return []

    def get_adapter_info(self) -> Dict[str, Any]:
        """
        Return metadata about this adapter for UI display and filtering.
        """
        return {
            "id": "cei_excel_format_v1",
            "name": "CEI Excel Format v1.2",
            "description": "Imports course, faculty, and section data from CEI's Excel exports. Supports original format (Faculty Name, effterm_c), test format (email, Term, students), and hybrid format (email, Term, Enrolled Students).",
            "supported_formats": [XLSX_EXTENSION, XLS_EXTENSION],
            "institution_short_name": "CEI",  # Identify by stable short name instead of GUID
            "data_types": ["courses", "faculty", "terms", "sections"],
            "version": "1.2.0",
            "created_by": "System Developer",
            "last_updated": "2024-09-25",
            "file_size_limit": "500MB",
            "max_records": 500000,
            "format_variants": [
                "Original: course, Faculty Name, effterm_c, students",
                "Test: course, email, Term, students",
                "Hybrid: course, email, Term, Enrolled Students",
            ],
        }

    def get_file_size_limit(self) -> int:
        """Get the file size limit in bytes for CEI files."""
        return self.MAX_FILE_SIZE_MB * 1024 * 1024

    def get_max_records(self) -> int:
        """Get the maximum number of records this adapter can process."""
        return self.MAX_RECORDS_TO_PROCESS

    def parse_file(
        self, file_path: str, options: Dict[str, Any]
    ) -> Dict[str, List[Dict]]:
        """
        Parse CEI Excel file into structured data ready for database import.

        Args:
            file_path: Path to the Excel file
            options: Import options including institution_id

        Returns:
            Dict mapping data types to lists of records
        """
        try:
            # Validate file and extract institution ID
            institution_id = self._validate_parse_inputs(file_path, options)

            # Read and validate Excel data
            df = self._read_excel_file(file_path)

            # Process all rows and collect entities
            result = self._process_excel_rows(df, institution_id)

            # Remove duplicates and validate results
            result = self._finalize_results(result)

            return result

        except (FileCompatibilityError, ValueError):
            raise
        except Exception as e:
            raise FileCompatibilityError(f"Error parsing file: {str(e)}") from e

    def _validate_parse_inputs(self, file_path: str, options: Dict[str, Any]) -> str:
        """Validate file compatibility and extract institution ID."""
        # Validate file first
        is_compatible, compatibility_message = self.validate_file_compatibility(
            file_path
        )
        if not is_compatible:
            raise FileCompatibilityError(f"File incompatible: {compatibility_message}")

        # Get institution ID from options
        institution_id = options.get("institution_id")
        if not institution_id:
            raise ValueError("institution_id is required in options")

        return institution_id

    def _read_excel_file(self, file_path: str) -> pd.DataFrame:
        """Read and validate Excel file."""
        try:
            df = pd.read_excel(file_path)
        except Exception as e:
            raise FileCompatibilityError(f"Cannot read Excel file: {str(e)}") from e

        if df.empty:
            raise FileCompatibilityError("Excel file is empty")

        return df

    def _process_excel_rows(
        self, df: pd.DataFrame, institution_id: str
    ) -> Dict[str, List[Dict]]:
        """Process all Excel rows and collect entities by type."""
        # Initialize result structure
        result: Dict[str, List[Dict]] = {
            "courses": [],
            "users": [],  # Note: using 'users' not 'faculty' to match database service
            "terms": [],
            "offerings": [],
            "sections": [],
        }

        # Process each row using existing parsing logic
        for _, row in df.iterrows():
            try:
                # Use existing parse_cei_excel_row function
                entities = parse_cei_excel_row(row, institution_id)

                # Add entities to result with timestamps
                self._collect_row_entities(entities, result)

            except Exception as e:
                # Log error but continue processing other rows
                print(f"Warning: Error processing row {_}: {str(e)}")
                continue

        return result

    def _collect_row_entities(
        self, entities: Dict[str, Any], result: Dict[str, List[Dict]]
    ) -> None:
        """Collect entities from a single row into the result structure."""
        timestamp = datetime.now().isoformat()

        # Map entity types to result keys
        entity_mappings = [
            ("course", "courses"),
            ("user", "users"),
            ("term", "terms"),
            ("offering", "offerings"),
            ("section", "sections"),
        ]

        for entity_key, result_key in entity_mappings:
            if entities.get(entity_key):
                entity = entities[entity_key].copy()
                entity["created_at"] = timestamp
                result[result_key].append(entity)

    def _finalize_results(self, result: Dict[str, List[Dict]]) -> Dict[str, List[Dict]]:
        """Remove duplicates and validate final results."""
        # Remove duplicates from each data type
        result = self._deduplicate_results(result)

        # Validate we have some data
        total_records = sum(len(records) for records in result.values())
        if total_records == 0:
            raise FileCompatibilityError("No valid records found in file")

        return result

    def _deduplicate_results(
        self, result: Dict[str, List[Dict]]
    ) -> Dict[str, List[Dict]]:
        """Remove duplicate records from parsed results."""

        # Define key fields for each data type to identify duplicates
        key_fields = {
            "courses": ["course_number", "institution_id"],
            "users": ["email", "first_name", "last_name", "institution_id"],
            "terms": ["name", "year", "institution_id"],
            "offerings": ["course_number", "term_name", "institution_id"],
            "sections": [
                "course_number",
                "term_name",
                "section_number",
                "institution_id",
            ],
        }

        for data_type, records in result.items():
            if not records:
                continue

            keys = key_fields.get(data_type, ["id"])
            seen = set()
            unique_records = []

            for record in records:
                # Create key tuple from specified fields
                key_values = []
                for field in keys:
                    value = record.get(field)
                    # Handle None values and convert to string for hashing
                    key_values.append(str(value) if value is not None else "")

                key = tuple(key_values)

                if key not in seen:
                    seen.add(key)
                    unique_records.append(record)

            result[data_type] = unique_records

        return result

    def export_data(
        self, data: Dict[str, List[Dict]], output_path: str, options: Dict[str, Any]
    ) -> Tuple[bool, str, int]:
        """
        Export structured data to CEI Excel format.

        Args:
            data: Structured data from database
            output_path: Where to save the Excel file
            options: Export configuration options

        Returns:
            Tuple[bool, str, int]: (success, message, records_exported)
        """
        try:
            from datetime import datetime

            from openpyxl import Workbook

            # Create workbook
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "CEI Export Data"

            # Build CEI-specific records by combining the data
            cei_records = self._build_cei_export_records(data, options)

            if not cei_records:
                return False, "No valid records to export", 0

            # Write headers (CEI format)
            headers = [
                "course",
                "section",
                "effterm_c",
                "students",
                FACULTY_NAME_COLUMN,
                "email",  # Include email if available
            ]

            for col, header in enumerate(headers, 1):
                worksheet.cell(row=1, column=col, value=header)

            # Write data rows
            for row, record in enumerate(cei_records, 2):
                worksheet.cell(row=row, column=1, value=record.get("course", ""))
                worksheet.cell(row=row, column=2, value=record.get("section", ""))
                worksheet.cell(row=row, column=3, value=record.get("effterm_c", ""))
                worksheet.cell(row=row, column=4, value=record.get("students", ""))
                worksheet.cell(
                    row=row, column=5, value=record.get(FACULTY_NAME_COLUMN, "")
                )
                worksheet.cell(row=row, column=6, value=record.get("email", ""))

            # Auto-size columns
            for col in range(1, len(headers) + 1):
                worksheet.column_dimensions[
                    worksheet.cell(row=1, column=col).column_letter
                ].width = 15

            # Save the workbook
            workbook.save(output_path)

            return (
                True,
                f"Successfully exported {len(cei_records)} records",
                len(cei_records),
            )

        except Exception as e:
            return False, f"Export failed: {str(e)}", 0

    def _build_cei_export_records(
        self, data: Dict[str, List[Dict]], options: Dict[str, Any]
    ) -> List[Dict[str, Any]]:
        """
        Build CEI-format records from standardized database data.

        Args:
            data: Database data organized by type
            options: Export options

        Returns:
            List of CEI-formatted records
        """
        records = []

        # Get data collections
        courses = data.get("courses", [])
        users = data.get("users", [])
        terms = data.get("terms", [])
        offerings = data.get("offerings", [])

        # Build a lookup for instructors
        instructors = [user for user in users if user.get("role") == "instructor"]

        # If we have actual course offerings, use them
        if offerings:
            # Build lookups for existing offerings logic
            instructors_lookup = {user["user_id"]: user for user in instructors}
            terms_lookup = {term["term_id"]: term for term in terms}

            # Process each course offering
            for offering in offerings:
                course_number = offering.get("course_number", "")
                term_id = offering.get("term_id")
                instructor_id = offering.get("instructor_id")

                # Find the course details
                course = next(
                    (c for c in courses if c.get("course_number") == course_number), {}
                )

                # Find the instructor
                instructor = instructors_lookup.get(instructor_id, {})

                # Find the term
                term = terms_lookup.get(term_id, {})

                # Format term for CEI (e.g., "2024 Fall" -> "2024FA")
                term_formatted = self._format_term_for_cei_export(term)

                record = {
                    "course": course_number,
                    "section": offering.get("section_number", "01"),
                    "effterm_c": term_formatted,
                    "students": offering.get("enrollment_count", 0),
                    FACULTY_NAME_COLUMN: f"{instructor.get('first_name', '')} {instructor.get('last_name', '')}".strip(),
                    "email": instructor.get("email", ""),
                }

                records.append(record)
        else:
            # Synthesize course offerings from available data
            # Create one record per course-instructor-term combination
            for course in courses:
                for instructor in instructors:
                    # Use the first available term or create a default
                    term = terms[0] if terms else {}

                    # Format term for CEI (e.g., "2024 Fall" -> "2024FA")
                    term_formatted = self._format_term_for_cei_export(term)

                    record = {
                        "course": course.get("course_number", ""),
                        "section": "01",  # Default section
                        "effterm_c": term_formatted,
                        "students": 25,  # Default student count
                        FACULTY_NAME_COLUMN: f"{instructor.get('first_name', '')} {instructor.get('last_name', '')}".strip(),
                        "email": instructor.get("email", ""),
                    }

                    records.append(record)

        return records

    def _format_term_for_cei_export(self, term: Dict[str, Any]) -> str:
        """Format term data for CEI export (e.g., SP2024)."""
        if not term:
            return ""

        year = term.get("year", "")
        season = term.get("season", "")

        if year and season:
            # Convert season to CEI format (SP2024, not 2024SP)
            season_map = {"Spring": "SP", "Summer": "SU", "Fall": "FA", "Winter": "WI"}
            season_code = season_map.get(season, season[:2].upper())
            return f"{season_code}{year}"

        # Fall back to parsing the term name if structured fields aren't available
        term_name = term.get("name", "")
        if term_name:
            try:
                # Parse term names like "Fall 2025", "Spring 2024", etc.
                parts = term_name.split()
                if len(parts) >= 2:
                    season_name = parts[0]  # "Fall", "Spring", etc.
                    year_str = parts[1]  # "2025", "2024", etc.

                    # Convert season to CEI format
                    season_map = {
                        "Spring": "SP",
                        "Summer": "SU",
                        "Fall": "FA",
                        "Winter": "WI",
                    }
                    season_code = season_map.get(season_name, season_name[:2].upper())

                    return f"{season_code}{year_str}"
            except (ValueError, IndexError):
                pass

        # If all else fails, return the original term name
        return term_name
